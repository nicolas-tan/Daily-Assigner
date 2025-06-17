#!/usr/bin/env python3
"""
Hybrid Daily Bug Assignment Script
Downloads Google Sheets, processes as Excel locally, optionally uploads back
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import gspread
from google.oauth2 import service_account
import os
import logging
from datetime import datetime
import tempfile
import shutil

# Import the original script's class
from script import CQEToOursNewest

logger = logging.getLogger(__name__)


class GoogleSheetsDownloader:
    """Helper class to download Google Sheets as Excel files"""
    
    def __init__(self, credentials_file='credentials.json'):
        """Initialize Google Sheets client"""
        self.creds = service_account.Credentials.from_service_account_file(
            credentials_file,
            scopes=['https://www.googleapis.com/auth/spreadsheets.readonly',
                   'https://www.googleapis.com/auth/drive.readonly']
        )
        self.client = gspread.authorize(self.creds)
        
    def download_as_excel(self, spreadsheet_id: str, output_path: str):
        """
        Download a Google Sheet as an Excel file
        
        Args:
            spreadsheet_id: Google Sheets ID
            output_path: Where to save the Excel file
        """
        try:
            # Open the spreadsheet
            spreadsheet = self.client.open_by_key(spreadsheet_id)
            
            # Create a new Excel workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Copy each sheet
            for sheet in spreadsheet.worksheets():
                ws = wb.create_sheet(title=sheet.title)
                
                # Get all values
                values = sheet.get_all_values()
                
                # Write to Excel sheet
                for row_idx, row in enumerate(values, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                        
            # Save the workbook
            wb.save(output_path)
            logger.info(f"Downloaded Google Sheet to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error downloading Google Sheet: {e}")
            return False


class HybridBugAssigner(CQEToOursNewest):
    """
    Hybrid approach: Download from Google Sheets, process locally, optionally upload back
    """
    
    def __init__(self, target_spreadsheet_id: str, source_spreadsheet_id: str = None,
                 credentials_file: str = 'credentials.json', 
                 keep_local_files: bool = False):
        """
        Initialize hybrid bug assigner
        
        Args:
            target_spreadsheet_id: ID of target Google Sheet
            source_spreadsheet_id: ID of source Google Sheet
            credentials_file: Path to Google credentials
            keep_local_files: Whether to keep downloaded Excel files
        """
        self.target_spreadsheet_id = target_spreadsheet_id
        self.source_spreadsheet_id = source_spreadsheet_id
        self.credentials_file = credentials_file
        self.keep_local_files = keep_local_files
        
        # Create temp directory for Excel files
        self.temp_dir = tempfile.mkdtemp(prefix='bug_assigner_')
        
        # Paths for temporary Excel files
        self.target_excel_path = os.path.join(self.temp_dir, 'target_bugs.xlsx')
        self.source_excel_path = os.path.join(self.temp_dir, 'source_bugs.xlsx')
        
        # Initialize downloader
        self.downloader = GoogleSheetsDownloader(credentials_file)
        
        # Download the sheets
        self._download_sheets()
        
        # Initialize parent class with local Excel files
        super().__init__(self.target_excel_path, self.source_excel_path)
        
    def _download_sheets(self):
        """Download Google Sheets as Excel files"""
        logger.info("Downloading Google Sheets...")
        
        # Download target sheet
        if not self.downloader.download_as_excel(self.target_spreadsheet_id, 
                                                self.target_excel_path):
            raise Exception("Failed to download target spreadsheet")
            
        # Download source sheet if provided
        if self.source_spreadsheet_id:
            if not self.downloader.download_as_excel(self.source_spreadsheet_id,
                                                    self.source_excel_path):
                logger.warning("Failed to download source spreadsheet")
                
    def upload_to_google_sheets(self):
        """Upload the processed Excel file back to Google Sheets"""
        logger.info("Uploading results back to Google Sheets...")
        
        try:
            # Initialize Google Sheets client with write permissions
            creds = service_account.Credentials.from_service_account_file(
                self.credentials_file,
                scopes=['https://www.googleapis.com/auth/spreadsheets']
            )
            client = gspread.authorize(creds)
            
            # Open target spreadsheet
            spreadsheet = client.open_by_key(self.target_spreadsheet_id)
            
            # Load the processed Excel file
            wb = load_workbook(self.target_excel_path)
            
            # Update each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                try:
                    # Get or create sheet in Google Sheets
                    try:
                        gs_sheet = spreadsheet.worksheet(sheet_name)
                    except:
                        gs_sheet = spreadsheet.add_worksheet(title=sheet_name, 
                                                           rows=1000, cols=20)
                    
                    # Clear existing content
                    gs_sheet.clear()
                    
                    # Prepare data for batch update
                    values = []
                    for row in ws.iter_rows(values_only=True):
                        values.append([str(cell) if cell is not None else '' 
                                     for cell in row])
                    
                    # Update in batch
                    if values:
                        gs_sheet.update('A1', values)
                        
                    logger.info(f"Updated sheet: {sheet_name}")
                    
                except Exception as e:
                    logger.error(f"Error updating sheet {sheet_name}: {e}")
                    
            logger.info("Upload completed successfully")
            
        except Exception as e:
            logger.error(f"Error uploading to Google Sheets: {e}")
            raise
            
    def cleanup(self):
        """Clean up temporary files"""
        if not self.keep_local_files and os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
            logger.info("Cleaned up temporary files")
            
    def run_full_process_with_upload(self):
        """Run the full process and upload results back to Google Sheets"""
        try:
            # Run the original process
            self.run_full_process()
            
            # Upload results back to Google Sheets
            self.upload_to_google_sheets()
            
            # Add link to Google Sheet in email
            self.google_sheet_link = f"https://docs.google.com/spreadsheets/d/{self.target_spreadsheet_id}/edit"
            
        finally:
            # Clean up
            self.cleanup()


def main():
    """Main function for hybrid approach"""
    
    # Configuration from environment
    TARGET_SPREADSHEET_ID = os.getenv('TARGET_SPREADSHEET_ID')
    SOURCE_SPREADSHEET_ID = os.getenv('SOURCE_SPREADSHEET_ID')
    CREDENTIALS_FILE = os.getenv('GOOGLE_CREDENTIALS_FILE', 'credentials.json')
    KEEP_LOCAL = os.getenv('KEEP_LOCAL_FILES', 'false').lower() == 'true'
    
    if not TARGET_SPREADSHEET_ID:
        logger.error("TARGET_SPREADSHEET_ID not set in environment")
        return
        
    # Create hybrid assigner
    assigner = HybridBugAssigner(
        TARGET_SPREADSHEET_ID,
        SOURCE_SPREADSHEET_ID,
        CREDENTIALS_FILE,
        KEEP_LOCAL
    )
    
    # Run the process with upload
    assigner.run_full_process_with_upload()
    
    logger.info("Hybrid process completed successfully!")


if __name__ == "__main__":
    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    main() 