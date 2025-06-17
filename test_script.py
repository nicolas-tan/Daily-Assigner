#!/usr/bin/env python3
"""
Test script for Daily Bug Assignment Automation
Creates sample data and tests the automation process
"""

import pandas as pd
import os
from datetime import datetime
from script import CQEToOursNewest
import openpyxl
from openpyxl import Workbook

def create_test_excel():
    """Create a test Excel file with required sheets"""
    wb = Workbook()
    
    # Create required sheets
    wb.active.title = "Daily New"
    wb.create_sheet("GL")
    wb.create_sheet("NT")
    wb.create_sheet("PP")
    
    # Add headers to Daily New sheet
    daily_sheet = wb["Daily New"]
    headers = ["Assignment", "Bug ID", "Priority", "Title", "Failure Mode", "Status", "Assignee"]
    for col, header in enumerate(headers, 1):
        daily_sheet.cell(row=1, column=col).value = header
    
    # Save the workbook
    test_file = "test_custom-top-cqe-bugs-daily-assigner.xlsx"
    wb.save(test_file)
    print(f"Created test Excel file: {test_file}")
    return test_file

def create_sample_bug_data():
    """Create sample bug data for testing"""
    sample_bugs = pd.DataFrame({
        'Bug ID': ['BUG-001', 'BUG-002', 'BUG-003', 'BUG-004', 'BUG-005', 
                   'BUG-006', 'BUG-007', 'BUG-008', 'BUG-009', 'BUG-010'],
        'Priority': [1, 5, 3, 2, 10, 7, 4, 6, 8, 9],
        'Title': [
            'Graphics rendering issue',
            'Network connectivity problem',
            'Power consumption spike',
            'Display driver crash',
            'WiFi disconnection',
            'Battery drain issue',
            'GPU memory leak',
            'Ethernet port failure',
            'Performance degradation',
            'Render pipeline error'
        ],
        'Failure Mode': [
            'Graphics card fails to render 3D objects',
            'Network adapter loses connection randomly',
            'Power usage exceeds normal limits',
            'Display driver crashes on startup',
            'WiFi module disconnects frequently',
            'Battery drains faster than expected',
            'GPU memory usage increases over time',
            'Ethernet connection drops intermittently',
            'System performance degrades under load',
            'Rendering pipeline produces artifacts'
        ],
        'Status': ['Open'] * 10,
        'Assignee': [''] * 10
    })
    
    # Save to CSV
    csv_file = "test_cqe_bugs.csv"
    sample_bugs.to_csv(csv_file, index=False)
    print(f"Created sample bug data: {csv_file}")
    return csv_file, sample_bugs

def test_basic_functionality():
    """Test basic functionality of the bug assignment system"""
    print("\n" + "="*50)
    print("Testing Daily Bug Assignment Automation")
    print("="*50 + "\n")
    
    # Create test files
    excel_file = create_test_excel()
    csv_file, sample_data = create_sample_bug_data()
    
    try:
        # Initialize the assigner
        print("\n1. Initializing bug assigner...")
        assigner = CQEToOursNewest(excel_file, csv_file)
        
        # Test loading workbook
        print("\n2. Loading workbook...")
        assigner.load_workbook()
        print("   ✓ Workbook loaded successfully")
        
        # Test grabbing CQE data
        print("\n3. Grabbing CQE data...")
        assigner.grab_CQE_daily(sample_data)
        print("   ✓ CQE data imported successfully")
        
        # Test assigning dropdown tags
        print("\n4. Assigning team tags...")
        assigner.assign_dropdown_tags()
        print("   ✓ Team assignments completed")
        
        # Test reordering by priority
        print("\n5. Reordering by priority...")
        assigner.reorder_by_priority('Daily New')
        print("   ✓ Bugs reordered by priority")
        
        # Test distributing to team sheets
        print("\n6. Distributing to team sheets...")
        assigner.distribute_to_team_sheets()
        print("   ✓ Bugs distributed to teams")
        
        # Test reordering team sheets
        print("\n7. Reordering team sheets...")
        for team in ['GL', 'NT', 'PP']:
            assigner.reorder_by_priority(team)
        print("   ✓ Team sheets reordered")
        
        # Test email preparation (without sending)
        print("\n8. Preparing email content...")
        assigner.send_daily_email()  # Will just prepare content without credentials
        print("   ✓ Email content prepared")
        
        print("\n" + "="*50)
        print("All tests passed successfully!")
        print("="*50)
        
        # Display summary
        print("\nSummary of bug assignments:")
        wb = openpyxl.load_workbook(excel_file)
        for sheet_name in ['GL', 'NT', 'PP']:
            sheet = wb[sheet_name]
            bug_count = sheet.max_row - 1  # Subtract header row
            print(f"  {sheet_name}: {bug_count} bugs assigned")
        
        print(f"\nTest files created:")
        print(f"  - Excel: {excel_file}")
        print(f"  - CSV: {csv_file}")
        print("\nYou can open these files to verify the results.")
        
    except Exception as e:
        print(f"\n❌ Error during testing: {e}")
        import traceback
        traceback.print_exc()
    
    # Cleanup option
    print("\nDo you want to keep the test files? (y/n): ", end="")
    response = input().strip().lower()
    if response != 'y':
        try:
            os.remove(excel_file)
            os.remove(csv_file)
            print("Test files removed.")
        except:
            pass

def test_assignment_logic():
    """Test the bug assignment logic"""
    print("\nTesting assignment logic...")
    
    test_cases = [
        ("Graphics rendering issue", "GL"),
        ("Network connectivity problem", "NT"),
        ("Power consumption spike", "PP"),
        ("GPU memory leak", "GL"),
        ("WiFi disconnection", "NT"),
        ("Battery drain issue", "PP"),
        ("Unknown issue type", "GL"),  # Default
    ]
    
    # Create a minimal instance just for testing
    assigner = CQEToOursNewest("dummy.xlsx")
    
    print("\nFailure Mode -> Team Assignment:")
    for failure_mode, expected in test_cases:
        result = assigner._determine_assignment(failure_mode)
        status = "✓" if result == expected else "✗"
        print(f"  {status} '{failure_mode}' -> {result} (expected: {expected})")

if __name__ == "__main__":
    # Run assignment logic test
    test_assignment_logic()
    
    # Run full functionality test
    print("\nDo you want to run the full functionality test? (y/n): ", end="")
    if input().strip().lower() == 'y':
        test_basic_functionality() 