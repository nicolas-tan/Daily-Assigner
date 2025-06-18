#!/usr/bin/env python3
"""
Streamlit App for Excel Bug Assignment Processor
"""

import streamlit as st
import pandas as pd
import os
import tempfile
from datetime import datetime
import io
from offline_processor import (
    is_cqe_file, 
    process_single_cqe_file,
    run_offline_process,
    create_blank_excel_template,
    create_team_sheets_email_html
)

# Set page config
st.set_page_config(
    page_title="Bug Assignment Processor",
    page_icon="🐛",
    layout="wide"
)

# Title and description
st.title("🐛 Daily Bug Assignment Processor")
st.markdown("""
This tool processes CQE bug files and assigns them to teams (GL, NT, PP) based on failure modes.
Upload your Excel file below to get started.
""")

# Create tabs
tab1, tab2, tab3 = st.tabs(["Process Files", "Create Template", "Help"])

with tab1:
    st.header("Upload and Process Bug Files")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xls'],
        help="Upload either a CQE bugs file or a properly formatted bug assignment file"
    )
    
    if uploaded_file is not None:
        # Save uploaded file to temp location
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            temp_path = tmp_file.name
        
        try:
            # Check if it's a CQE file
            if is_cqe_file(temp_path):
                st.info("📊 Detected CQE file format - will create proper structure and process")
                
                if st.button("Process CQE File", type="primary"):
                    with st.spinner("Processing CQE file..."):
                        # Process the file
                        output_file = process_single_cqe_file(temp_path)
                        
                        if output_file and os.path.exists(output_file):
                            st.success("✅ Processing completed successfully!")
                            
                            # Read the processed file
                            with open(output_file, 'rb') as f:
                                processed_data = f.read()
                            
                            # Offer download
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.download_button(
                                    label="📥 Download Processed Excel File",
                                    data=processed_data,
                                    file_name=os.path.basename(output_file),
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            
                            with col2:
                                # Generate HTML report
                                html_content = create_team_sheets_email_html(output_file)
                                st.download_button(
                                    label="📄 Download HTML Report",
                                    data=html_content,
                                    file_name=f"bug_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
                                    mime="text/html"
                                )
                            
                            # Show summary
                            st.subheader("📊 Processing Summary")
                            
                            # Read the Excel file to show summary
                            import openpyxl
                            wb = openpyxl.load_workbook(output_file)
                            
                            summary_data = []
                            for team in ['GL', 'NT', 'PP']:
                                ws = wb[team]
                                count = 0
                                for row in range(2, ws.max_row + 1):
                                    if any(ws.cell(row=row, column=col).value for col in range(1, 9)):
                                        count += 1
                                summary_data.append({'Team': team, 'Bug Count': count})
                            
                            wb.close()
                            
                            # Display summary
                            summary_df = pd.DataFrame(summary_data)
                            col1, col2, col3 = st.columns(3)
                            
                            for idx, row in summary_df.iterrows():
                                with [col1, col2, col3][idx]:
                                    st.metric(row['Team'], f"{row['Bug Count']} bugs")
                            
                            st.metric("Total Bugs", f"{summary_df['Bug Count'].sum()} bugs")
                            
                            # Clean up
                            os.remove(output_file)
                        else:
                            st.error("❌ Processing failed. Please check your file format.")
            
            else:
                st.info("📋 Detected properly formatted bug assignment file")
                st.warning("⚠️ Note: Processing files with existing structure is not yet implemented in the Streamlit version")
                
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)

with tab2:
    st.header("Create Blank Template")
    st.markdown("""
    Create a blank Excel template with the required structure for bug assignments.
    This template includes sheets for Daily New, GL, NT, and PP teams.
    """)
    
    if st.button("Generate Blank Template", type="primary"):
        with st.spinner("Creating template..."):
            # Create template in temp location
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                template_path = tmp_file.name
            
            # Generate the template
            create_blank_excel_template(template_path)
            
            # Read the template
            with open(template_path, 'rb') as f:
                template_data = f.read()
            
            # Offer download
            st.download_button(
                label="📥 Download Blank Template",
                data=template_data,
                file_name="blank_bug_assignment_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.success("✅ Template created successfully!")
            
            # Clean up
            os.remove(template_path)

with tab3:
    st.header("How to Use")
    
    st.markdown("""
    ### 📋 File Types Supported
    
    1. **CQE Bug Files**: Excel files from the CQE system containing bug data
    2. **Bug Assignment Files**: Pre-formatted files with the proper structure
    
    ### 🔄 Processing Steps
    
    1. **Upload** your Excel file using the file uploader
    2. **Click Process** to run the bug assignment algorithm
    3. **Download** the processed Excel file and/or HTML report
    
    ### 📊 Assignment Logic
    
    - **Power/Thermal Issues** → Assigned to PP team
    - **Other Issues** → Alternately assigned to GL and NT teams
    - **Non-SXM5 Products** → Highlighted in black and excluded from team sheets
    
    ### 🏷️ Failure Mode Mapping
    
    The system automatically maps failure modes based on keywords in bug titles:
    - OVERT → Thermal
    - SLD, POWER → Power
    - L2, CACHE, RMINIT → SRAM Memory
    - REMAP → HBM Memory
    - And more...
    """)
    
    st.info("💡 Tip: The HTML report can be opened directly in a browser or shared with team members")

# Footer
st.markdown("---")
st.markdown("Built with Streamlit 🚀") 