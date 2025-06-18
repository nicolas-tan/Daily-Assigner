#!/usr/bin/env python3
"""
Offline Excel Processing Script
Run the bug assignment operations on manually downloaded Excel files
"""

import os
import sys
import logging
from datetime import datetime
from script import CQEToOursNewest
import pandas as pd

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def is_cqe_file(excel_path):
    """
    Check if the given Excel file is a CQE bugs file (doesn't have our required sheets)
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        # Check if it has our required sheets
        required_sheets = ['Daily New', 'GL', 'NT', 'PP']
        has_required_sheets = all(sheet in sheet_names for sheet in required_sheets)
        
        # If it has the required sheets BUT also has CQE-specific sheets, it's still a CQE file
        cqe_indicators = ['From External CSP', 'Closed cases', 'Pivot by TAT', 'StatusLocation']
        has_cqe_sheets = any(any(indicator in sheet for indicator in cqe_indicators) for sheet in sheet_names)
        
        return has_cqe_sheets or not has_required_sheets
    except Exception:
        return True  # Assume it's a CQE file if we can't read it properly


def process_single_cqe_file(cqe_file_path):
    """
    Process a single CQE file by creating a proper structure and importing data from first tab only
    """
    logger.info(f"Processing CQE file: {cqe_file_path}")
    
    # Create output filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = os.path.splitext(os.path.basename(cqe_file_path))[0]
    output_file = f"processed_{timestamp}_{base_name}.xlsx"
    
    # Step 1: Read ONLY the first sheet of CQE data
    logger.info("Reading CQE data from first sheet only...")
    try:
        # Get sheet names first
        excel_file = pd.ExcelFile(cqe_file_path)
        first_sheet_name = excel_file.sheet_names[0]
        logger.info(f"Reading from sheet: '{first_sheet_name}'")
        
        # Read only the first sheet
        cqe_data = pd.read_excel(cqe_file_path, sheet_name=0)
        logger.info(f"Found {len(cqe_data)} rows in first sheet")
        logger.info(f"Columns: {', '.join(cqe_data.columns.tolist())}")
    except Exception as e:
        logger.error(f"Error reading CQE file: {e}")
        return None
    
    # Step 2: Map CQE columns to our format
    logger.info("Mapping CQE columns to standard format...")
    
    # Define column mapping based on user requirements
    column_mapping = {
        'SXM SN for SXM RMA': 'SN Associated',
        'FA NVbug Number': 'Bug ID',
        'Priority': 'Priority',
        'Customer Reported Failure': 'Title',
        'Date added to Top 30 of priority list': 'Created Date',
        'Status': 'COMPLETED',
        'Product': 'Product'
    }
    
    # Create a new DataFrame with mapped columns
    mapped_data = pd.DataFrame()
    
    # Map each column
    for cqe_col, new_col in column_mapping.items():
        if cqe_col in cqe_data.columns:
            mapped_data[new_col] = cqe_data[cqe_col]
            logger.info(f"Mapped '{cqe_col}' to '{new_col}'")
        else:
            logger.warning(f"Column '{cqe_col}' not found in CQE data")
            mapped_data[new_col] = ''
    
    # Debug: Check Product column values
    if 'Product' in mapped_data.columns:
        logger.info(f"Product column values (first 10): {mapped_data['Product'].head(10).tolist()}")
    
    # Process Bug ID column - add https://nvbugs/ prefix if not present
    if 'Bug ID' in mapped_data.columns:
        def format_bug_id(bug_id):
            if pd.isna(bug_id) or str(bug_id).strip() == '':
                return 'nvbug not provided for this bug.'
            bug_id_str = str(bug_id).strip()
            # Check if it already has the prefix
            if bug_id_str.startswith('https://nvbugs/'):
                return bug_id_str
            # If it's just a number, add the prefix
            return f'https://nvbugs/{bug_id_str}'
        
        mapped_data['Bug ID'] = mapped_data['Bug ID'].apply(format_bug_id)
        logger.info("Formatted Bug ID column with https://nvbugs/ prefix")
    
    # Convert SN Associated column to handle large numbers
    if 'SN Associated' in mapped_data.columns:
        # Convert to numeric first to clean the data
        mapped_data['SN Associated'] = pd.to_numeric(mapped_data['SN Associated'], errors='coerce')
        # Fill NaN values with 0
        mapped_data['SN Associated'] = mapped_data['SN Associated'].fillna(0)
        # Convert large numbers to integers then to string to preserve full number
        # This prevents scientific notation display
        mapped_data['SN Associated'] = mapped_data['SN Associated'].apply(
            lambda x: str(int(x)) if x != 0 and pd.notna(x) else ''
        )
    
    # Add empty columns for missing fields
    mapped_data['Assignment'] = ''  # Empty Assignment column
    mapped_data['Failure Mode'] = ''  # Empty as requested
    
    # Map Failure Mode based on Title keywords
    if 'Title' in mapped_data.columns:
        logger.info("Mapping Failure Mode based on Title keywords...")
        
        def map_failure_mode(title):
            if pd.isna(title) or str(title).strip() == '':
                return 'Unknown Cause Issue'  # Changed from '' to 'Unknown Cause Issue' for blank titles
            
            title_upper = str(title).upper()
            
            # Check keywords in order of priority
            if 'OVERT' in title_upper:
                return 'Thermal'
            elif 'SLD' in title_upper or 'POWER' in title_upper:
                return 'Power'
            elif 'L2' in title_upper or 'CACHE' in title_upper or 'RMINIT' in title_upper:
                return 'SRAM Memory'
            elif 'REMAP' in title_upper:
                return 'HBM Memory'
            elif 'ECC' in title_upper or 'UEC' in title_upper or ' CE ' in title_upper or 'DATA MISMATCH' in title_upper or 'CRC' in title_upper or 'DATA/BUFFER' in title_upper:
                return 'General Memory (SRAM or HBM)'
            elif 'SUDDEN DEATH' in title_upper or 'SDC' in title_upper or 'THERMAL' in title_upper:
                return 'Thermal'
            elif 'XID' in title_upper:
                return 'Customer ID Specific Issue'
            elif 'FALLING OFF' in title_upper:
                return 'GPU Falling Off Bus'
            elif 'DEVICE INTERRUPT' in title_upper or 'INTERRUPT' in title_upper or 'IST' in title_upper:
                return 'Unknown Cause Issue'
            elif 'PCIE' in title_upper:
                return 'Potential Memory or NVSpec Issues'
            else:
                return 'Unknown Cause Issue'  # Changed from '' to 'Unknown Cause Issue' for unmatched cases
        
        # Apply the mapping
        mapped_data['Failure Mode'] = mapped_data['Title'].apply(map_failure_mode)
        
        # Log some examples
        failure_mode_counts = mapped_data['Failure Mode'].value_counts()
        logger.info("Failure Mode mapping results:")
        for mode, count in failure_mode_counts.items():
            if mode:  # Only log non-empty modes
                logger.info(f"  {mode}: {count} bugs")
    
    # Track rows that need black highlighting (non-SXM5 products) - BEFORE reordering columns
    rows_to_highlight = []
    if 'Product' in mapped_data.columns:
        logger.info(f"Product column found, checking {len(mapped_data)} rows for highlighting...")
        for idx, product in enumerate(mapped_data['Product']):
            if pd.notna(product) and str(product).strip() != '':
                product_str = str(product).strip()
                # Check if "SXM5" is NOT in the product string
                if 'SXM5' not in product_str:
                    rows_to_highlight.append(idx)
                    logger.info(f"Row {idx+2} will be highlighted (Product: {product_str})")
            else:
                # Empty product values should also be highlighted
                rows_to_highlight.append(idx)
                logger.debug(f"Row {idx+2} has empty/NaN Product value - will be highlighted")
    else:
        logger.warning("Product column not found in mapped_data!")
    
    # Reorder columns to match our template - now including Product as rightmost
    column_order = ['Assignment', 'Bug ID', 'Priority', 'Title', 'Failure Mode', 
                    'Created Date', 'COMPLETED', 'SN Associated', 'Product']
    
    # Ensure all columns exist
    for col in column_order:
        if col not in mapped_data.columns:
            mapped_data[col] = ''
    
    # Don't remove Product column anymore - keep it as rightmost column
    # if 'Product' in mapped_data.columns:
    #     del mapped_data['Product']
    
    mapped_data = mapped_data[column_order]
    
    # Process Priority column - fill blanks with sequential numbers
    if 'Priority' in mapped_data.columns:
        priority_col = mapped_data['Priority'].copy()
        
        # Convert to numeric, keeping NaN for blanks
        priority_col = pd.to_numeric(priority_col, errors='coerce')
        
        # Fill blanks with sequential numbers
        last_value = 0
        for i in range(len(priority_col)):
            if pd.notna(priority_col.iloc[i]):
                # Found a non-blank value
                current_value = int(priority_col.iloc[i])
                if i > 0 and last_value > 0:
                    # Fill the blanks between last_value and current_value
                    blank_start = None
                    for j in range(i-1, -1, -1):
                        if pd.isna(priority_col.iloc[j]):
                            blank_start = j
                        else:
                            break
                    
                    if blank_start is not None:
                        # Fill sequential values
                        num_blanks = i - blank_start
                        if current_value > last_value + num_blanks:
                            # Fill with sequential numbers
                            for k, idx in enumerate(range(blank_start, i)):
                                priority_col.iloc[idx] = last_value + k + 1
                        else:
                            # Just fill with incremental values
                            for k, idx in enumerate(range(blank_start, i)):
                                priority_col.iloc[idx] = last_value + k + 1
                
                last_value = current_value
            elif i == 0:
                # First row is blank, start with 1
                priority_col.iloc[i] = 1
                last_value = 1
        
        # Handle any remaining blanks at the end
        for i in range(len(priority_col)):
            if pd.isna(priority_col.iloc[i]):
                priority_col.iloc[i] = last_value + 1
                last_value += 1
        
        # Update the mapped_data with filled priorities
        mapped_data['Priority'] = priority_col.astype(int)
        logger.info("Filled blank Priority values with sequential numbers")
    
    # Assign teams based on Failure Mode
    logger.info("Assigning teams based on Failure Mode...")
    gl_nt_toggle = True  # Start with GL
    
    for idx in range(len(mapped_data)):
        failure_mode = mapped_data.loc[idx, 'Failure Mode']
        
        if failure_mode in ['Power', 'Thermal']:
            mapped_data.loc[idx, 'Assignment'] = 'PP'
        else:
            # Alternate between GL and NT for non-Power/Thermal bugs
            mapped_data.loc[idx, 'Assignment'] = 'GL' if gl_nt_toggle else 'NT'
            gl_nt_toggle = not gl_nt_toggle  # Toggle for next non-PP bug
    
    # Log assignment summary
    assignment_counts = mapped_data['Assignment'].value_counts()
    logger.info("Team assignment results:")
    for team, count in assignment_counts.items():
        logger.info(f"  {team}: {count} bugs")
    
    # Step 3: Create output Excel with proper structure
    logger.info("Creating output Excel file...")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.styles.colors import Color
    from openpyxl.worksheet.datavalidation import DataValidation
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create required sheets
    sheets = ['Daily New', 'GL', 'NT', 'PP']
    
    # Define black fill for non-SXM5 rows - using explicit opaque black
    black_color = Color(rgb="FF000000", type="rgb")
    black_fill = PatternFill(patternType="solid", fgColor=black_color, bgColor=black_color)
    white_font = Font(color="FFFFFF", bold=True)  # White text on black background
    
    for sheet_name in sheets:
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        for col_idx, header in enumerate(column_order, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # Format header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # If this is the Daily New sheet, add all the data
        if sheet_name == 'Daily New':
            # Create data validation for Assignment column (column 1)
            dv = DataValidation(type="list", formula1='"GL,NT,PP"', allow_blank=True)
            dv.error = 'Please select a valid team'
            dv.errorTitle = 'Invalid Team'
            dv.prompt = 'Please select a team from the list'
            dv.promptTitle = 'Team Assignment'
            
            # Add the data validation to the worksheet
            ws.add_data_validation(dv)
            
            for row_idx, row_data in mapped_data.iterrows():
                excel_row = row_idx + 2  # Excel row number (1-indexed, plus header)
                
                # Check if this row should be highlighted
                should_highlight = row_idx in rows_to_highlight
                
                for col_idx, (col_name, value) in enumerate(zip(column_order, row_data), 1):
                    cell = ws.cell(row=excel_row, column=col_idx, value=value)
                    
                    # Add dropdown to Assignment column (column 1)
                    if col_idx == 1:  # Assignment column
                        dv.add(cell)
                    
                    # Apply black highlighting if needed
                    if should_highlight:
                        cell.fill = black_fill
                        cell.font = white_font
                    
                    # Special formatting for SN Associated column (last column)
                    if col_name == 'SN Associated' and value and value != '':
                        # Store as text to preserve full number without scientific notation
                        cell.value = str(value)
                        # Format to look like a number (right-aligned)
                        cell.alignment = Alignment(horizontal='right')
                        if should_highlight:
                            cell.fill = black_fill
                            cell.font = white_font
        
        # Adjust column widths
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
    
    # Save the workbook
    wb.save(output_file)
    logger.info(f"Created output file with {len(mapped_data)} bugs")
    
    # Debug: Verify highlighting was applied
    logger.info("Verifying highlighting in saved file...")
    wb_check = openpyxl.load_workbook(output_file)
    ws_check = wb_check['Daily New']
    for idx in rows_to_highlight[:5]:  # Check first 5 highlighted rows
        cell = ws_check.cell(row=idx+2, column=1)
        logger.info(f"Row {idx+2}: Fill = {cell.fill.fgColor.rgb if cell.fill.fgColor else 'None'}")
    wb_check.close()
    
    # Step 4: Process with the standard workflow
    # NOTE: We'll only do the distribution part to preserve highlighting
    logger.info("Distributing bugs to team sheets based on Assignment...")
    
    # Load the workbook we just created
    wb_dist = openpyxl.load_workbook(output_file)
    
    # Get the Daily New sheet
    daily_new_sheet = wb_dist['Daily New']
    
    # Clear existing data in team sheets (except headers)
    for team in ['GL', 'NT', 'PP']:
        team_sheet = wb_dist[team]
        # Delete all rows except header
        team_sheet.delete_rows(2, team_sheet.max_row)
    
    # Track next available row for each team
    team_next_row = {'GL': 2, 'NT': 2, 'PP': 2}
    
    # Iterate through Daily New sheet and copy rows to appropriate team sheets
    for row in range(2, daily_new_sheet.max_row + 1):
        assignment = daily_new_sheet.cell(row=row, column=1).value  # Assignment is column 1
        
        if assignment in ['GL', 'NT', 'PP']:
            # Get the target sheet
            target_sheet = wb_dist[assignment]
            target_row = team_next_row[assignment]
            
            # Copy the entire row including formatting
            for col in range(1, daily_new_sheet.max_column + 1):
                source_cell = daily_new_sheet.cell(row=row, column=col)
                target_cell = target_sheet.cell(row=target_row, column=col)
                
                # Copy value
                target_cell.value = source_cell.value
                
                # Copy formatting (including black highlighting)
                if source_cell.fill.patternType:
                    target_cell.fill = PatternFill(
                        patternType=source_cell.fill.patternType,
                        fgColor=source_cell.fill.fgColor,
                        bgColor=source_cell.fill.bgColor
                    )
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        color=source_cell.font.color
                    )
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical
                    )
            
            # Add dropdown to Assignment column in team sheet
            dv_team = DataValidation(type="list", formula1='"GL,NT,PP"', allow_blank=True)
            dv_team.error = 'Please select a valid team'
            dv_team.errorTitle = 'Invalid Team'
            dv_team.prompt = 'Please select a team from the list'
            dv_team.promptTitle = 'Team Assignment'
            target_sheet.add_data_validation(dv_team)
            dv_team.add(target_sheet.cell(row=target_row, column=1))
            
            # Increment the row counter for this team
            team_next_row[assignment] += 1
    
    # Log distribution results
    logger.info("Distribution results:")
    for team in ['GL', 'NT', 'PP']:
        count = team_next_row[team] - 2  # Subtract 2 to get actual count (header + starting at 2)
        logger.info(f"  {team}: {count} bugs")
    
    # Save the workbook with distributed data
    wb_dist.save(output_file)
    wb_dist.close()
    
    # Step 5: Remove black-highlighted rows from team sheets
    logger.info("Removing black-highlighted rows from team sheets...")
    
    # Reload the workbook to clean up black rows
    wb_cleanup = openpyxl.load_workbook(output_file)
    
    for team in ['GL', 'NT', 'PP']:
        team_sheet = wb_cleanup[team]
        rows_to_delete = []
        
        # Check each row for black highlighting (starting from row 2 to skip header)
        for row in range(2, team_sheet.max_row + 1):
            # Check if any cell in the row has black fill
            cell = team_sheet.cell(row=row, column=1)  # Check first cell
            if cell.fill.fgColor and cell.fill.fgColor.rgb == 'FF000000':
                rows_to_delete.append(row)
        
        # Delete rows in reverse order to maintain correct row numbers
        for row in reversed(rows_to_delete):
            team_sheet.delete_rows(row, 1)
            logger.info(f"Removed black-highlighted row from {team} sheet")
        
        if rows_to_delete:
            logger.info(f"Removed {len(rows_to_delete)} black-highlighted rows from {team} sheet")
    
    # Save the cleaned workbook
    wb_cleanup.save(output_file)
    wb_cleanup.close()
    
    # Log final distribution after cleanup
    logger.info("Final distribution after removing black-highlighted rows:")
    wb_final = openpyxl.load_workbook(output_file)
    for team in ['GL', 'NT', 'PP']:
        team_sheet = wb_final[team]
        # Count non-empty rows (excluding header)
        count = 0
        for row in range(2, team_sheet.max_row + 1):
            if any(team_sheet.cell(row=row, column=col).value for col in range(1, 9)):
                count += 1
        logger.info(f"  {team}: {count} bugs")
    wb_final.close()
    
    logger.info("=== Process completed successfully! ===")
    logger.info(f"Output saved to: {output_file}")
    
    # Step 6: Create and send email with team sheet contents
    logger.info("Preparing bug assignment report...")
    
    # Create HTML email content
    html_content = create_team_sheets_email_html(output_file)
    
    # Save HTML report
    report_filename = f"bug_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    with open(report_filename, 'w') as f:
        f.write(html_content)
    
    logger.info(f"\n{'='*60}")
    logger.info(f"HTML Report saved to: {report_filename}")
    logger.info(f"You can:")
    logger.info(f"  1. Open it directly in a browser")
    logger.info(f"  2. Host it with: python -m http.server 8080")
    logger.info(f"  3. Upload it to any web hosting service")
    logger.info(f"{'='*60}\n")
    
    # Start HTTP server to serve the report
    serve_html_report(report_filename)
    
    return output_file


def create_team_sheets_email_html(excel_file):
    """Create HTML content with tables for each team sheet"""
    import openpyxl
    from datetime import datetime
    
    wb = openpyxl.load_workbook(excel_file)
    
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            h1 {{ color: #333; }}
            h2 {{ color: #666; margin-top: 30px; }}
            table {{ border-collapse: collapse; margin-bottom: 30px; width: 100%; }}
            th {{ background-color: #4CAF50; color: white; padding: 12px; text-align: left; }}
            td {{ border: 1px solid #ddd; padding: 8px; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .summary {{ background-color: #e7f3ff; padding: 15px; margin-bottom: 20px; border-radius: 5px; }}
        </style>
    </head>
    <body>
        <h1>Daily Bug Assignment Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}</h1>
        <div class="summary">
            <strong>Summary:</strong><br>
    """
    
    # Add summary counts
    team_counts = {}
    for team in ['GL', 'NT', 'PP']:
        ws = wb[team]
        count = 0
        for row in range(2, ws.max_row + 1):
            if any(ws.cell(row=row, column=col).value for col in range(1, 9)):
                count += 1
        team_counts[team] = count
        html += f"&nbsp;&nbsp;&nbsp;&nbsp;{team}: {count} bugs<br>"
    
    html += f"&nbsp;&nbsp;&nbsp;&nbsp;<strong>Total: {sum(team_counts.values())} bugs</strong>"
    html += "</div>"
    
    # Create table for each team
    for team in ['GL', 'NT', 'PP']:
        ws = wb[team]
        html += f"<h2>{team} Team - {team_counts[team]} bugs</h2>"
        html += "<table>"
        
        # Add headers
        html += "<tr>"
        headers = ['Assignment', 'Bug ID', 'Priority', 'Title', 'Failure Mode', 
                  'Created Date', 'COMPLETED', 'SN Associated', 'Product']
        for header in headers:
            html += f"<th>{header}</th>"
        html += "</tr>"
        
        # Add data rows
        for row in range(2, ws.max_row + 1):
            # Check if row has data
            if any(ws.cell(row=row, column=col).value for col in range(1, 9)):
                html += "<tr>"
                for col in range(1, 10):  # 9 columns including Product
                    value = ws.cell(row=row, column=col).value or ""
                    # Truncate long titles for better display
                    if col == 4 and len(str(value)) > 50:  # Title column
                        value = str(value)[:50] + "..."
                    html += f"<td>{value}</td>"
                html += "</tr>"
        
        html += "</table>"
    
    html += """
    </body>
    </html>
    """
    
    wb.close()
    return html


def run_offline_process(target_excel_path, source_excel_path=None):
    """
    Run the Excel processing operations offline
    
    Args:
        target_excel_path: Path to your custom-top-cqe-bugs-daily-assigner.xlsx file
        source_excel_path: Path to the CQE source data file (optional)
    """
    logger.info("=== Starting Offline Excel Processing ===")
    logger.info(f"Target file: {target_excel_path}")
    if source_excel_path:
        logger.info(f"Source file: {source_excel_path}")
    
    # Check if files exist
    if not os.path.exists(target_excel_path):
        logger.error(f"Target file not found: {target_excel_path}")
        return False
        
    if source_excel_path and not os.path.exists(source_excel_path):
        logger.error(f"Source file not found: {source_excel_path}")
        return False
    
    try:
        # Create processor instance
        processor = CQEToOursNewest(target_excel_path, source_excel_path)
        
        # Run the full process
        logger.info("Running full bug assignment process...")
        processor.run_full_process()
        
        logger.info("=== Process completed successfully! ===")
        logger.info(f"Updated file saved at: {target_excel_path}")
        
        # Create a backup of the processed file
        backup_name = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(target_excel_path)}"
        backup_path = os.path.join(os.path.dirname(target_excel_path), backup_name)
        
        # Copy the file as backup
        import shutil
        shutil.copy2(target_excel_path, backup_path)
        logger.info(f"Backup saved at: {backup_path}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error during processing: {e}")
        import traceback
        traceback.print_exc()
        return False


def create_blank_excel_template(output_path="blank_template.xlsx"):
    """
    Create a blank Excel template with the required sheets and structure
    """
    import openpyxl
    
    logger.info(f"Creating blank Excel template at: {output_path}")
    
    # Create new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create required sheets
    sheets = ['Daily New', 'GL', 'NT', 'PP']
    
    # Define headers for each sheet - updated to remove Status column
    headers = [
        'Assignment', 'Bug ID', 'Priority', 'Title', 'Failure Mode', 
        'Created Date', 'COMPLETED', 'SN Associated'
    ]
    
    for sheet_name in sheets:
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            
        # Format header row
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            
        # Adjust column widths
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
    
    # Save the workbook
    wb.save(output_path)
    logger.info(f"Blank template created: {output_path}")
    
    return output_path


def main():
    """Main function for offline processing"""
    
    print("\n=== Offline Excel Bug Assignment Processor ===\n")
    
    # Check command line arguments
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python offline_processor.py <excel_file>")
        print("\nExamples:")
        print("  python offline_processor.py cqe_bugs.xlsx")
        print("  python offline_processor.py custom-top-cqe-bugs-daily-assigner.xlsx cqe_bugs.xlsx")
        print("\nTo create a blank template:")
        print("  python offline_processor.py --create-template")
        sys.exit(1)
    
    # Check if creating template
    if sys.argv[1] == "--create-template":
        template_name = sys.argv[2] if len(sys.argv) > 2 else "blank_template.xlsx"
        create_blank_excel_template(template_name)
        sys.exit(0)
    
    # Get file paths
    first_file = sys.argv[1]
    
    # Check if file exists
    if not os.path.exists(first_file):
        print(f"❌ Error: File not found: {first_file}")
        sys.exit(1)
    
    # Determine if single file or two file mode
    if len(sys.argv) == 2:
        # Single file mode - check if it's a CQE file
        if is_cqe_file(first_file):
            logger.info("Detected CQE file format - will create proper structure")
            output_file = process_single_cqe_file(first_file)
            if output_file:
                print(f"\n✅ Processing completed successfully!")
                print(f"📄 Output file: {output_file}")
                print(f"🔗 You can now upload this file to Google Sheets or Excel Online")
            else:
                print("\n❌ Processing failed. Check the logs above for details.")
                sys.exit(1)
        else:
            # It already has proper structure, process it alone
            success = run_offline_process(first_file)
            if success:
                print("\n✅ Processing completed successfully!")
            else:
                print("\n❌ Processing failed. Check the logs above for details.")
                sys.exit(1)
    else:
        # Two file mode - traditional target + source
        target_file = first_file
        source_file = sys.argv[2]
        
        success = run_offline_process(target_file, source_file)
        
        if success:
            print("\n✅ Processing completed successfully!")
        else:
            print("\n❌ Processing failed. Check the logs above for details.")
            sys.exit(1)


def kill_port_8080():
    """Kill any process using port 8080"""
    import subprocess
    try:
        # Find process using the port
        result = subprocess.run(['lsof', '-t', '-i', ':8080'], 
                              capture_output=True, text=True)
        if result.stdout.strip():
            pids = result.stdout.strip().split('\n')
            for pid in pids:
                logger.info(f"Killing existing process {pid} on port 8080")
                subprocess.run(['kill', pid])
            # Give it a moment to release the port
            import time
            time.sleep(1)
            logger.info("Port 8080 cleared")
    except Exception as e:
        logger.debug(f"Note: Could not check for existing processes: {e}")


def serve_html_report(report_filename):
    """Start HTTP server to serve the HTML report"""
    import http.server
    import socketserver
    import socket
    import threading
    
    PORT = 8080
    
    # Kill any existing server on port 8080
    kill_port_8080()
    
    # Get network info
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip_address = s.getsockname()[0]
        s.close()
    except:
        ip_address = "localhost"
    
    print(f"\n{'='*70}")
    print(f"🌐 HTML Report is now being served!")
    print(f"{'='*70}")
    print(f"📄 Report: {report_filename}")
    print(f"\n🔗 Access the report at:")
    print(f"   Local:    http://localhost:{PORT}/{report_filename}")
    print(f"   Network:  http://{ip_address}:{PORT}/{report_filename}")
    print(f"{'='*70}")
    print(f"Press Ctrl+C to stop the server and exit\n")
    
    # Start server
    Handler = http.server.SimpleHTTPRequestHandler
    try:
        with socketserver.TCPServer(("", PORT), Handler) as httpd:
            httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n✅ Server stopped. Goodbye!")
    except Exception as e:
        print(f"\n❌ Error starting server: {e}")
        print(f"You can manually serve the report with: python -m http.server {PORT}")


if __name__ == "__main__":
    main() 