#!/usr/bin/env python3
"""
Daily Bug Assignment Automation Script
Automates the process of scraping CQE bugs, assigning them to team members,
and sending daily email notifications.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
import os
import sys
from typing import Dict, List, Tuple
import logging
import re

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class CQEToOursNewest:
    """Main class for automating daily bug assignment process"""
    
    def __init__(self, excel_path: str, cqe_source_path: str = None):
        """
        Initialize the bug assignment automation
        
        Args:
            excel_path: Path to the custom-top-cqe-bugs-daily-assigner.xlsx
            cqe_source_path: Path to the CQE source data (optional)
        """
        self.excel_path = excel_path
        self.cqe_source_path = cqe_source_path
        self.workbook = None
        self.sheets = {
            'Daily New': None,
            'GL': None,
            'NT': None,
            'PP': None
        }
        self.team_emails = {
            'GL': 'geetha@company.com',
            'NT': 'nicolas@company.com',
            'PP': 'phuong@company.com'
        }
        
    def load_workbook(self):
        """Load the Excel workbook and verify required sheets exist"""
        try:
            self.workbook = load_workbook(self.excel_path, data_only=False)
            
            # Verify all required sheets exist
            required_sheets = ['Daily New', 'GL', 'NT', 'PP']
            existing_sheets = self.workbook.sheetnames
            
            for sheet in required_sheets:
                if sheet not in existing_sheets:
                    logger.warning(f"Sheet '{sheet}' not found. Creating it...")
                    self.workbook.create_sheet(sheet)
                    
            # Load sheet references
            for sheet_name in self.sheets.keys():
                self.sheets[sheet_name] = self.workbook[sheet_name]
                
            logger.info("Workbook loaded successfully")
            
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            raise
            
    def grab_CQE_daily(self, source_data: pd.DataFrame = None):
        """
        Step 1 & 2: Grab CQE data and paste into Daily New sheet
        
        Args:
            source_data: DataFrame containing CQE bug data (if not provided, will try to read from file)
        """
        logger.info("Starting CQE data grab...")
        
        if source_data is None and self.cqe_source_path:
            # Read from file if path provided
            try:
                if self.cqe_source_path.endswith('.xlsx'):
                    source_data = pd.read_excel(self.cqe_source_path)
                elif self.cqe_source_path.endswith('.csv'):
                    source_data = pd.read_csv(self.cqe_source_path)
                else:
                    logger.error("Unsupported file format")
                    return
            except Exception as e:
                logger.error(f"Error reading source data: {e}")
                return
                
        if source_data is None:
            logger.warning("No source data provided")
            return
            
        # Get existing data from Daily New sheet
        daily_sheet = self.sheets['Daily New']
        existing_data = self._sheet_to_dataframe(daily_sheet)
        
        # Process new bugs
        for _, new_bug in source_data.iterrows():
            bug_id = new_bug.get('Bug ID', new_bug.get('ID', ''))
            priority = new_bug.get('Priority', 100)
            
            # Check if bug already exists
            if not existing_data.empty and bug_id in existing_data.get('Bug ID', []).values:
                # Update priority if bug exists
                row_idx = existing_data[existing_data['Bug ID'] == bug_id].index[0]
                daily_sheet.cell(row=row_idx + 2, column=self._get_column_index(daily_sheet, 'Priority')).value = priority
                logger.info(f"Updated priority for existing bug {bug_id}")
            else:
                # Add new bug
                self._append_bug_to_sheet(daily_sheet, new_bug)
                logger.info(f"Added new bug {bug_id}")
                
        self.workbook.save(self.excel_path)
        
    def assign_dropdown_tags(self):
        """
        Step 3: Assign dropdown tags to new bugs based on failure mode
        """
        logger.info("Assigning dropdown tags...")
        
        daily_sheet = self.sheets['Daily New']
        
        # Create data validation for dropdown
        dv = DataValidation(
            type="list",
            formula1='"GL,NT,PP"',
            allow_blank=True
        )
        dv.error = 'Please select GL, NT, or PP'
        dv.errorTitle = 'Invalid Entry'
        
        # Find the assignment column (usually column A)
        assignment_col = 1
        
        # Apply to all data rows
        max_row = daily_sheet.max_row
        for row in range(2, max_row + 1):  # Skip header row
            cell = daily_sheet.cell(row=row, column=assignment_col)
            
            # Auto-assign based on failure mode if not already assigned
            if not cell.value:
                failure_mode = daily_sheet.cell(row=row, column=self._get_column_index(daily_sheet, 'Failure Mode')).value
                assignment = self._determine_assignment(failure_mode)
                cell.value = assignment
                
            # Add dropdown validation
            dv.add(cell)
            
        daily_sheet.add_data_validation(dv)
        self.workbook.save(self.excel_path)
        
    def _determine_assignment(self, failure_mode: str) -> str:
        """
        Determine team assignment based on failure mode
        
        Args:
            failure_mode: The failure mode description
            
        Returns:
            Team assignment (GL, NT, or PP)
        """
        if not failure_mode:
            return ""
            
        failure_mode = str(failure_mode).lower()
        
        # Define assignment rules based on failure mode patterns
        gl_patterns = ['graphics', 'display', 'render', 'gpu']
        nt_patterns = ['network', 'connectivity', 'wifi', 'ethernet']
        pp_patterns = ['power', 'performance', 'battery', 'thermal']
        
        for pattern in gl_patterns:
            if pattern in failure_mode:
                return "GL"
                
        for pattern in nt_patterns:
            if pattern in failure_mode:
                return "NT"
                
        for pattern in pp_patterns:
            if pattern in failure_mode:
                return "PP"
                
        # Default assignment logic
        return "GL"  # Default to GL if no pattern matches
        
    def reorder_by_priority(self, sheet_name: str = 'Daily New'):
        """
        Step 4 & 6: Reorder bugs by priority
        
        Args:
            sheet_name: Name of the sheet to reorder
        """
        logger.info(f"Reordering bugs by priority in {sheet_name}...")
        
        sheet = self.sheets[sheet_name]
        data = self._sheet_to_dataframe(sheet)
        
        if data.empty:
            return
            
        # Sort by priority
        data_sorted = data.sort_values('Priority', ascending=True)
        
        # Clear sheet except header
        self._clear_sheet_data(sheet)
        
        # Write sorted data back
        for idx, row in data_sorted.iterrows():
            row_num = idx + 2  # Account for header
            for col_idx, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_idx).value = value
                
        self.workbook.save(self.excel_path)
        
    def distribute_to_team_sheets(self):
        """
        Step 5: Distribute bugs to respective team sheets (GL, NT, PP)
        """
        logger.info("Distributing bugs to team sheets...")
        
        daily_sheet = self.sheets['Daily New']
        
        # Get all bugs with assignments
        for row in range(2, daily_sheet.max_row + 1):
            assignment = daily_sheet.cell(row=row, column=1).value
            
            if assignment in ['GL', 'NT', 'PP']:
                # Copy entire row to respective sheet
                row_data = []
                for col in range(1, daily_sheet.max_column + 1):
                    row_data.append(daily_sheet.cell(row=row, column=col).value)
                    
                team_sheet = self.sheets[assignment]
                self._append_row_to_sheet(team_sheet, row_data)
                
        self.workbook.save(self.excel_path)
        
    def delete_completed_bugs(self):
        """
        Step 7: Delete completed bugs (green highlighted cells)
        """
        logger.info("Deleting completed bugs...")
        
        for sheet_name in self.sheets.keys():
            sheet = self.sheets[sheet_name]
            
            # Add COMPLETED column if it doesn't exist
            completed_col = self._get_or_create_column(sheet, 'COMPLETED')
            
            # Check for completed bugs (green fill)
            rows_to_delete = []
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=completed_col)
                
                # Check if cell has green fill
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == "FF00FF00":
                    rows_to_delete.append(row)
                    
            # Delete rows in reverse order to maintain indices
            for row in reversed(rows_to_delete):
                sheet.delete_rows(row)
                logger.info(f"Deleted completed bug in row {row} from {sheet_name}")
                
        self.workbook.save(self.excel_path)
        
    def send_daily_email(self, smtp_server: str = "smtp.gmail.com", smtp_port: int = 587,
                        sender_email: str = None, sender_password: str = None):
        """
        Step 8: Send daily email with top 25 priority bugs for each team
        
        Args:
            smtp_server: SMTP server address
            smtp_port: SMTP server port
            sender_email: Sender's email address
            sender_password: Sender's email password
        """
        logger.info("Preparing daily email...")
        
        # Prepare email content
        subject = f"Daily Top Bugs Ready For You to Start - {datetime.now().strftime('%Y-%m-%d')}"
        
        body = """
        <html>
        <body>
        <h2>Daily Top Bugs Assignment</h2>
        <p>Please review your assigned bugs below. You can click on rows you don't want and click to 'empty' 
        (on manual click, bug will stay only in Tab#1)</p>
        """
        
        # Get top 25 bugs for each team
        for team in ['GL', 'NT', 'PP']:
            sheet = self.sheets[team]
            data = self._sheet_to_dataframe(sheet)
            
            if not data.empty:
                # Get top 25 by priority
                top_bugs = data.nsmallest(25, 'Priority')
                
                body += f"<h3>{team} Team Bugs:</h3>"
                body += "<table border='1' style='border-collapse: collapse;'>"
                body += "<tr><th>Bug ID</th><th>Priority</th><th>Title</th><th>Failure Mode</th></tr>"
                
                for _, bug in top_bugs.iterrows():
                    body += f"<tr>"
                    body += f"<td>{bug.get('Bug ID', '')}</td>"
                    body += f"<td>{bug.get('Priority', '')}</td>"
                    body += f"<td>{bug.get('Title', '')}</td>"
                    body += f"<td>{bug.get('Failure Mode', '')}</td>"
                    body += f"</tr>"
                    
                body += "</table><br>"
                
        body += """
        </body>
        </html>
        """
        
        # Send email
        if sender_email and sender_password:
            try:
                msg = MIMEMultipart('alternative')
                msg['Subject'] = subject
                msg['From'] = sender_email
                msg['To'] = ', '.join(self.team_emails.values())
                
                msg.attach(MIMEText(body, 'html'))
                
                with smtplib.SMTP(smtp_server, smtp_port) as server:
                    server.starttls()
                    server.login(sender_email, sender_password)
                    server.send_message(msg)
                    
                logger.info("Email sent successfully")
                
            except Exception as e:
                logger.error(f"Error sending email: {e}")
        else:
            logger.warning("Email credentials not provided. Skipping email send.")
            logger.info(f"Email content prepared:\n{body}")
            
    def run_full_process(self, source_data: pd.DataFrame = None):
        """
        Run the complete daily bug assignment process
        
        Args:
            source_data: DataFrame containing new CQE bug data
        """
        logger.info("Starting full daily bug assignment process...")
        
        try:
            # Load workbook
            self.load_workbook()
            
            # Step 1 & 2: Grab and paste CQE data
            self.grab_CQE_daily(source_data)
            
            # Step 3: Assign dropdown tags
            self.assign_dropdown_tags()
            
            # Step 4: Reorder by priority in Daily New
            self.reorder_by_priority('Daily New')
            
            # Step 5: Distribute to team sheets
            self.distribute_to_team_sheets()
            
            # Step 6: Reorder by priority in team sheets
            for team in ['GL', 'NT', 'PP']:
                self.reorder_by_priority(team)
                
            # Step 7: Delete completed bugs
            self.delete_completed_bugs()
            
            # Step 8: Send email
            # Note: You'll need to provide email credentials
            self.send_daily_email()
            
            logger.info("Daily bug assignment process completed successfully!")
            
        except Exception as e:
            logger.error(f"Error in daily process: {e}")
            raise
            
    # Helper methods
    def _sheet_to_dataframe(self, sheet) -> pd.DataFrame:
        """Convert worksheet to pandas DataFrame"""
        data = []
        headers = []
        
        # Get headers from first row
        for col in range(1, sheet.max_column + 1):
            headers.append(sheet.cell(row=1, column=col).value or f"Column{col}")
            
        # Get data
        for row in range(2, sheet.max_row + 1):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                row_data.append(sheet.cell(row=row, column=col).value)
            data.append(row_data)
            
        return pd.DataFrame(data, columns=headers)
        
    def _get_column_index(self, sheet, column_name: str) -> int:
        """Get column index by name"""
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                return col
        return None
        
    def _clear_sheet_data(self, sheet):
        """Clear all data except header row"""
        for row in range(sheet.max_row, 1, -1):
            sheet.delete_rows(row)
            
    def _append_bug_to_sheet(self, sheet, bug_data):
        """Append bug data to sheet"""
        row_num = sheet.max_row + 1
        col_num = 2  # Start from column B
        
        # Standard column mapping
        columns = ['Bug ID', 'Priority', 'Title', 'Failure Mode', 'Status', 'Assignee']
        
        for col in columns:
            if col in bug_data:
                sheet.cell(row=row_num, column=col_num).value = bug_data[col]
            col_num += 1
            
    def _append_row_to_sheet(self, sheet, row_data: List):
        """Append a row of data to sheet"""
        row_num = sheet.max_row + 1
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_idx).value = value
            
    def _get_or_create_column(self, sheet, column_name: str) -> int:
        """Get column index or create if doesn't exist"""
        col_idx = self._get_column_index(sheet, column_name)
        
        if col_idx is None:
            # Add new column
            col_idx = sheet.max_column + 1
            sheet.cell(row=1, column=col_idx).value = column_name
            
        return col_idx


def main():
    """Main function to run the daily bug assignment"""
    
    # Configuration
    EXCEL_PATH = "custom-top-cqe-bugs-daily-assigner.xlsx"
    CQE_SOURCE_PATH = "cqe_bugs.csv"  # Or path to CQE data source
    
    # Email configuration (set these environment variables or update directly)
    SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
    SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
    SENDER_EMAIL = os.getenv('SENDER_EMAIL')
    SENDER_PASSWORD = os.getenv('SENDER_PASSWORD')
    
    # Create instance
    assigner = CQEToOursNewest(EXCEL_PATH, CQE_SOURCE_PATH)
    
    # Run the full process
    # Note: You can also pass a DataFrame directly if you have the data in memory
    assigner.run_full_process()
    
    # For email functionality, ensure credentials are set:
    # assigner.send_daily_email(SMTP_SERVER, SMTP_PORT, SENDER_EMAIL, SENDER_PASSWORD)


if __name__ == "__main__":
    main()
